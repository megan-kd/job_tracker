from flask import Flask, Response
from openpyxl import load_workbook
import datetime

# globals
DEFAULT_EXCEL_PATH = "data/jobappdata.xlsx"
DEFAULT_COMPANY_POSITION_COL = 0 #A
DEFAULT_APP_LINK_COL = 5 #F
DEFAULT_DATE_APPLIED_COL = 6 #G
DEFAULT_DATE_REJECTED_COL = 7 #H
DEFAULT_STATUS_COL = 3
DEFAULT_INTERVIEW_COL = 2 #C
CURRENT_SHEET_NAME = "Current"
REJECTED_SHEET_NAME = "Rejected"
DATETIME_FORMAT = "%Y-%m-%d 00:00:00"
VALID_APP_TYPES = {"rejected", "current"}
REJECTED_STATUS = "Rejected"
JOB_GONE_STATUS = "Job No Longer Exists"
ACTIVE_STATUS = "Active"
WITHDRAWN_STATUS = "Withdrawn"

wb = load_workbook(DEFAULT_EXCEL_PATH)
current_ws = wb[CURRENT_SHEET_NAME]
rejected_ws = wb[REJECTED_SHEET_NAME]

def parse_apps(sheet_type):
    if sheet_type == "current":
        worksheet = current_ws
    elif sheet_type == "rejected":
        worksheet = rejected_ws

    current_apps_list = []
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        current_app = {}
        # if role and row is empty, don't include
        if row[DEFAULT_COMPANY_POSITION_COL].value is None:
            continue
        else:
            role = str(row[DEFAULT_COMPANY_POSITION_COL].value).split(" - ")
            current_app["Company"] = role[0]
            current_app["Position Title"] = role[1]
            current_app["Date Applied"] = datetime.datetime.strptime(str(row[DEFAULT_DATE_APPLIED_COL].value), DATETIME_FORMAT)
            current_app["Link"] = str(row[DEFAULT_APP_LINK_COL].value)
            current_app["Status"] = ACTIVE_STATUS
            if sheet_type == "rejected":
                if str(row[DEFAULT_STATUS_COL].value) == REJECTED_STATUS:
                    current_app["Date Rejected"] = datetime.datetime.strptime(str(row[DEFAULT_DATE_REJECTED_COL].value), DATETIME_FORMAT)
                    current_app["Status"] = REJECTED_STATUS
                    current_app["Response Time"] = int((current_app["Date Rejected"] - current_app["Date Applied"]).total_seconds() / 86400)
                elif str(row[DEFAULT_STATUS_COL].value) == JOB_GONE_STATUS:
                    current_app["Date Rejected"] = "None"
                    current_app["Status"] = JOB_GONE_STATUS
                elif  str(row[DEFAULT_STATUS_COL].value) == WITHDRAWN_STATUS:
                    current_app["Status"] = WITHDRAWN_STATUS
                if row[DEFAULT_INTERVIEW_COL].value is None:
                    current_app["Interview"] = "No"
                else:
                    current_app["Interview"] = str(row[DEFAULT_INTERVIEW_COL].value)
            current_apps_list.append(current_app)
    return {"data": current_apps_list}


#TODOS
# (stretch) pull excel from google sheets with API maybe?
# Chart Data to Get
    # Histogram -> Buckets of Response Times


all_current_apps = parse_apps("current")
all_rejected_apps = parse_apps("rejected")


app = Flask(__name__)


@app.route("/")
def hello_world():
    return "<p>Hello World!</p>"


# make dict of all current job apps 
#   (Company, Position, Date Applied, Link to Job Description)
# make dict of all rejected job apps
#   (Company, Position, Date Applied, Date Rejected, Link to Job Description)
@app.route("/get_all_apps/<sheet>")
def get_all_apps(sheet):
    if sheet == "current":
        return all_current_apps
    elif sheet == "rejected":
        return all_rejected_apps
    elif sheet == "all":
        return {"current": all_current_apps, "rejected": all_rejected_apps }
    else:
        return Response("Invalid Sheet Type", status=400, mimetype='application/json')

# ------------------------------ CHART DATA ENDPOINTS -----------------------------

# Pie Chart 
    # -> Total Applications (Rejected, In the Running)
    # -> Total Applications (Interviews, No Interview, Ghosted) 
@app.route("/pie/<dataset>")
def get_total_application_pie_data(dataset):
    if dataset == "status":
        return {"data": [
            {"type": "rejected", "value": len(all_rejected_apps["data"])},
            {"type": "waiting", "value": len(all_current_apps["data"])},
        ]}
    elif dataset == "responses":
        interviewed = 0
        not_interviewed = 0
        ghosted = 0

        # parse rejected apps
        for application in all_rejected_apps["data"]:
            if application["Status"] == JOB_GONE_STATUS:
                ghosted += 1
            elif application["Interview"] != "No":
                interviewed += 1
            else:
                not_interviewed += 1

        return {"data": [
            {"type": "interviewed", "value": interviewed},
            {"type": "no interview", "value": not_interviewed},
            {"type": "ghosted", "value": ghosted},
        ]}
    else:
        return Response("Invalid Pie Data Type", status=400, mimetype='application/json')


# Table of companies response time
@app.route("/response_time_data")
def response_time_data():
    data = {"active": [], "rejected": []}
    now = datetime.datetime.now()
    # waiting application use current date to find length
    for application in all_current_apps["data"]:
        data["active"].append({
            "Company": application["Company"],
            "Response Time": int((now - application["Date Applied"]).total_seconds() / 86400),
            "Status": application["Status"]
        })

    # only go through ones that are rejected and get the response time (days)
    for application in all_rejected_apps["data"]:
        if application["Status"] == REJECTED_STATUS:
            data["rejected"].append({
                "Company": application["Company"],
                "Response Time": application["Response Time"],
                "Status": application["Status"],
            })
    return {"data": data}

# Number of Applications that needed a job account made: 
    # Workday Accounts
    # ICIMS Accounts or Ulti (UKPRO) or SuccessFactors
@app.route("/job_account_data")
def job_account_data():
    data = {
        "iCIMS": 0,
        "Workday": 0,
        "SuccessFactors": 0,
        "UKGPRO": 0
    }
    ICIMS_KEYWORD = "icims"
    WORKDAY_KEYWORD = "workday"
    SF_KEYWORD = "successfactors"
    ULTIPRO_KEYWORD = "ultipro"

    for application in all_current_apps["data"]:
        if ICIMS_KEYWORD in application["Link"]:
            data["iCIMS"] += 1
        elif WORKDAY_KEYWORD in application["Link"]:
            data["Workday"] += 1
        elif SF_KEYWORD in application["Link"]:
            data["SuccessFactors"] += 1
        elif ULTIPRO_KEYWORD in application["Link"]:
            data["UKGPRO"] += 1

    for application in all_rejected_apps["data"]:
        if ICIMS_KEYWORD in application["Link"]:
            data["iCIMS"] += 1
        elif WORKDAY_KEYWORD in application["Link"]:
            data["Workday"] += 1
        elif SF_KEYWORD in application["Link"]:
            data["SuccessFactors"] += 1
        elif ULTIPRO_KEYWORD in application["Link"]:
            data["UKGPRO"] += 1

    return {"data": data}


if __name__ == '__main__':
    app.run(debug=True, port=8000)