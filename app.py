from flask import Flask, Response
from openpyxl import load_workbook
import datetime

# globals
DEFAULT_EXCEL_PATH = "data/jobappdata.xlsx"
DEFAULT_COMPANY_POSITION_COL = 0 #A
DEFAULT_APP_LINK_COL = 5 #F
DEFAULT_DATE_APPLIED_COL = 6 #G
DEFAULT_DATE_REJECTED_COL = 7 #H
DEFAULT_INTERVIEW_COL = 2 #C
CURRENT_SHEET_NAME = "Current"
REJECTED_SHEET_NAME = "Rejected"
DATETIME_FORMAT = "%Y-%m-%d 00:00:00"
VALID_APP_TYPES = {"rejected", "current"}

wb = load_workbook(DEFAULT_EXCEL_PATH)
current_ws = wb[CURRENT_SHEET_NAME]
rejected_ws = wb[REJECTED_SHEET_NAME]

#TODOS
# (stretch) pull excel from google sheets with API maybe?
# get total count of current jobs
# get total count of rejected jobs
# Chart Data to Get
    # Pie Chart -> Total Applications (Rejected, In the Running)
    # Pie Chart -> Total Applications (Interviews, No Interview, Haven't Heard Back)
    # Histogram -> Buckets of Response Times
    # Table -> Sort list of companies by avg response time

def parse_apps(sheet_type):
    worksheet = None
    if sheet_type not in VALID_APP_TYPES:
        return Response("Invalid Sheet Type", status=400, mimetype='application/json')
    elif sheet_type == "current":
        worksheet = current_ws
    elif sheet_type == "rejected":
        worksheet = rejected_ws

    current_apps_list = []
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        current_app = {}
        # if role and row is empty, don't include
        if row[DEFAULT_COMPANY_POSITION_COL].value is None:
            continue
        else:
            role = str(row[DEFAULT_COMPANY_POSITION_COL].value).split(" - ")
            current_app["Company"] = role[0]
            current_app["Position Title"] = role[1]
            current_app["Date Applied"] = datetime.datetime.strptime(str(row[DEFAULT_DATE_APPLIED_COL].value), DATETIME_FORMAT)
            current_app["Link"] = str(row[DEFAULT_APP_LINK_COL].value)
            if sheet_type == "rejected":
                current_app["Date Rejected"] = datetime.datetime.strptime(str(row[DEFAULT_DATE_REJECTED_COL].value), DATETIME_FORMAT)
            current_apps_list.append(current_app)
    return {sheet_type: current_apps_list}



app = Flask(__name__)


@app.route("/")
def hello_world():
    return "<p>Hello World!</p>"


# make dictionary of all current job apps (Company, Position Title, Date Applied, Link to Job Description)
# make dictionary of all rejected job apps (Company, Position Title, Date Applied, Date, Rejected, Link to Job Description)
@app.route("/get_all_apps/<sheet>")
def get_apps_in_the_running(sheet):
    return parse_apps(sheet)


